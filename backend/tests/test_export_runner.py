"""导出 Worker 的领取、完成与精确补偿行为。"""

import asyncio
import inspect
import io
from datetime import UTC, datetime
from pathlib import Path
from typing import cast
from uuid import UUID

import pytest
from openpyxl import load_workbook

from app.export.models import ClaimedExport, FrozenExport
from app.export.tasks import ExportRetryRequired, ExportRunner
from app.export.worker_repository import SqlAlchemyExportWorkerRepository
from app.storage.supabase import SupabaseObjectStorage, SupabaseStorageError
from tests.test_export_xlsx import batch, rows

EXPORT_ID = UUID("11111111-1111-4111-8111-111111111111")


def test_export_lease_outlives_the_worker_hard_time_limit() -> None:
    parameter = inspect.signature(SqlAlchemyExportWorkerRepository.claim).parameters[
        "lease_seconds"
    ]
    assert parameter.default == 600
    assert parameter.default > 570


class Repository:
    def __init__(
        self,
        *,
        complete: bool = True,
        fail: bool = True,
        claim: bool = True,
        retry_after_seconds: int | None = None,
    ) -> None:
        self.should_complete = complete
        self.should_fail = fail
        self.should_claim = claim
        self.retry_after_seconds = retry_after_seconds
        self.completed = 0
        self.failed: list[str] = []

    async def claim(self, export_id: UUID, lease_token: UUID) -> ClaimedExport | None:
        if not self.should_claim:
            return None
        return ClaimedExport(
            export=FrozenExport(
                id=export_id,
                export_type="draft",
                workbook_schema_version="paper-grading-workbook.v1",
                snapshot_at=datetime(2026, 7, 22, tzinfo=UTC),
                generation_started_at=datetime(2026, 7, 22, 0, 1, tzinfo=UTC),
                batch_snapshot=batch(1),
                items=rows(1),
            ),
            lease_token=lease_token,
        )

    async def claim_retry_delay_seconds(self, _export_id: UUID) -> int | None:
        return self.retry_after_seconds

    async def complete(self, *_args: object, **_kwargs: object) -> bool:
        self.completed += 1
        return self.should_complete

    async def fail(self, _export_id: UUID, _lease_token: UUID, code: str) -> bool:
        self.failed.append(code)
        return self.should_fail


class Storage:
    def __init__(
        self,
        *,
        created: bool = True,
        delete_fails: bool = False,
        upload_fails: bool = False,
    ) -> None:
        self.created = created
        self.delete_fails = delete_fails
        self.upload_fails = upload_fails
        self.uploaded: bytes | None = None
        self.deleted = False

    async def put_file_once(
        self,
        _key: str,
        path: Path,
        **_kwargs: object,
    ) -> bool:
        if self.upload_fails:
            raise SupabaseStorageError("upload failed")
        self.uploaded = path.read_bytes()
        return self.created

    async def delete_if_sha256(self, _key: str, _sha256: bytes) -> bool:
        if self.delete_fails:
            raise SupabaseStorageError("cleanup failed")
        self.deleted = True
        return True


def test_export_runner_completes_from_frozen_snapshot() -> None:
    repository = Repository()
    storage = Storage()
    runner = ExportRunner(
        repository=cast(SqlAlchemyExportWorkerRepository, repository),
        storage=cast(SupabaseObjectStorage, storage),
    )
    assert asyncio.run(runner.run(EXPORT_ID)) == "completed"
    assert repository.completed == 1
    assert storage.uploaded is not None and storage.uploaded.startswith(b"PK")
    workbook = load_workbook(io.BytesIO(storage.uploaded), data_only=False, keep_links=True)
    metadata = workbook["Metadata"]
    values = {
        metadata.cell(row, 1).value: metadata.cell(row, 2).value
        for row in range(2, metadata.max_row + 1)
    }
    assert values["snapshot_at"] == "2026-07-22T00:00:00+00:00"
    assert values["generated_at_utc"] == "2026-07-22T00:01:00+00:00"
    assert storage.deleted is False


def test_active_lease_requests_delayed_retry_instead_of_acknowledging_the_message() -> None:
    repository = Repository(claim=False, retry_after_seconds=321)
    runner = ExportRunner(
        repository=cast(SqlAlchemyExportWorkerRepository, repository),
        storage=cast(SupabaseObjectStorage, Storage()),
    )

    with pytest.raises(ExportRetryRequired) as retry:
        asyncio.run(runner.run(EXPORT_ID))

    assert retry.value.countdown_seconds == 321


def test_terminal_duplicate_is_acknowledged_without_retry() -> None:
    repository = Repository(claim=False, retry_after_seconds=None)
    runner = ExportRunner(
        repository=cast(SqlAlchemyExportWorkerRepository, repository),
        storage=cast(SupabaseObjectStorage, Storage()),
    )

    assert asyncio.run(runner.run(EXPORT_ID)) == "not_claimed"


def test_repeated_soft_timeout_is_failed_without_generating_another_workbook() -> None:
    repository = Repository()
    storage = Storage()
    runner = ExportRunner(
        repository=cast(SqlAlchemyExportWorkerRepository, repository),
        storage=cast(SupabaseObjectStorage, storage),
    )

    assert asyncio.run(runner.fail_timed_out(EXPORT_ID)) == "export_workbook_timeout"
    assert repository.failed == ["export_workbook_timeout"]
    assert storage.uploaded is None


def test_export_runner_deletes_a_new_object_only_after_its_lease_is_atomically_failed() -> None:
    repository = Repository(complete=False)
    storage = Storage()
    runner = ExportRunner(
        repository=cast(SqlAlchemyExportWorkerRepository, repository),
        storage=cast(SupabaseObjectStorage, storage),
    )
    assert asyncio.run(runner.run(EXPORT_ID)) == "export_completion_failed"
    assert repository.failed == ["export_completion_failed"]
    assert storage.deleted is True


def test_old_worker_never_deletes_after_a_new_lease_has_taken_ownership() -> None:
    repository = Repository(complete=False, fail=False)
    storage = Storage()
    runner = ExportRunner(
        repository=cast(SqlAlchemyExportWorkerRepository, repository),
        storage=cast(SupabaseObjectStorage, storage),
    )

    assert asyncio.run(runner.run(EXPORT_ID)) == "lease_lost"
    assert storage.deleted is False


def test_reclaimed_worker_reuses_an_identical_existing_object() -> None:
    repository = Repository()
    storage = Storage(created=False)
    runner = ExportRunner(
        repository=cast(SqlAlchemyExportWorkerRepository, repository),
        storage=cast(SupabaseObjectStorage, storage),
    )

    assert asyncio.run(runner.run(EXPORT_ID)) == "completed"
    assert repository.completed == 1
    assert storage.deleted is False


def test_storage_failure_marks_the_export_failed_without_completing() -> None:
    repository = Repository()
    runner = ExportRunner(
        repository=cast(SqlAlchemyExportWorkerRepository, repository),
        storage=cast(SupabaseObjectStorage, Storage(upload_fails=True)),
    )

    assert asyncio.run(runner.run(EXPORT_ID)) == "export_storage_failed"
    assert repository.failed == ["export_storage_failed"]
    assert repository.completed == 0


def test_cleanup_failure_crashes_without_relabeling_the_database_failure() -> None:
    repository = Repository(complete=False)
    storage = Storage(delete_fails=True)
    runner = ExportRunner(
        repository=cast(SqlAlchemyExportWorkerRepository, repository),
        storage=cast(SupabaseObjectStorage, storage),
    )

    with pytest.raises(RuntimeError, match="export_cleanup_failed"):
        asyncio.run(runner.run(EXPORT_ID))
    assert repository.failed == ["export_completion_failed"]
