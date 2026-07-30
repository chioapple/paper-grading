"""阶段十二四工作表与不可信文本边界。"""

import hashlib
import io
from datetime import UTC, datetime
from decimal import Decimal
from uuid import UUID

import pytest
from openpyxl import load_workbook

from app.export.xlsx import (
    WorkbookSnapshot,
    WorkbookValidationError,
    build_export_workbook,
    workbook_snapshot_from_frozen,
)

EXPORT_ID = UUID("11111111-1111-4111-8111-111111111111")


def batch(count: int) -> dict[str, object]:
    return {
        "schema_version": "export-batch-snapshot.v1",
        "assignment_id": "22222222-2222-4222-8222-222222222222",
        "grading_job_id": "33333333-3333-4333-8333-333333333333",
        "assignment_title": "Argumentative essay",
        "rubric_version_id": "44444444-4444-4444-8444-444444444444",
        "rubric_version": 1,
        "rubric": {
            "schema_version": 1,
            "total_score": "10",
            "score_step": "0.25",
            "dimensions": [
                {
                    "id": "argument",
                    "name": "Argument",
                    "description": "Quality of argument.",
                    "max_score": "10",
                    "bands": [
                        {
                            "label": "0",
                            "min_score": "0",
                            "max_score": "0",
                            "description": "Missing.",
                        },
                        {
                            "label": "1-10",
                            "min_score": "0.25",
                            "max_score": "10",
                            "description": "Present.",
                        },
                    ],
                    "evidence_requirements": ["Quote one block."],
                }
            ],
            "deductions": [
                {
                    "id": "late",
                    "name": "Late",
                    "description": "Late penalty.",
                    "points": "1",
                }
            ],
        },
        "provider_config_id": "55555555-5555-4555-8555-555555555555",
        "provider_config_version": 2,
        "model": "model-v1",
        "model_parameters_hash": "11" * 32,
        "prompt_version": "grading-prompt.v3",
        "prompt_hash": "22" * 32,
        "result_schema_version": "grade-result.v1",
        "result_schema_hash": "33" * 32,
        "rubric_hash": "44" * 32,
        "paper_count": count,
    }


def rows(count: int, *, filename: str = "essay.pdf") -> tuple[dict[str, object], ...]:
    return tuple(
        {
            "position": position,
            "submission_id": UUID(int=position + 10),
            "grading_attempt_id": UUID(int=position + 200),
            "teacher_review_id": None,
            "review_revision": None,
            "source_type": "ai_suggestion",
            "original_filename": filename if position == 0 else f"essay-{position}.pdf",
            "result_snapshot": {
                "schema_version": "export-item-snapshot.v1",
                "item_status": "needs_review",
                "max_score": 10,
                "subtotal": Decimal("9.25"),
                "deduction_total": Decimal("1"),
                "final_score": Decimal("8.25"),
                "criteria_results": [
                    {
                        "dimension_id": "argument",
                        "score": "9.25",
                        "reason": "Clear argument.",
                        "evidence": [{"block_id": "b000001", "quote": "Evidence"}],
                        "revision_suggestions": ["Add one counterargument."],
                    }
                ],
                "deduction_results": [
                    {
                        "deduction_id": "late",
                        "applied": True,
                        "reason": "Submitted late.",
                        "evidence": [],
                    }
                ],
                "evidence": [],
                "overall_feedback": "Strong work.",
                "change_reason": None,
                "confirmed_at": None,
            },
        }
        for position in range(count)
    )


def snapshot(count: int = 1, *, filename: str = "essay.pdf") -> WorkbookSnapshot:
    return workbook_snapshot_from_frozen(
        EXPORT_ID,
        "draft",
        "paper-grading-workbook.v1",
        datetime(2026, 7, 22, tzinfo=UTC),
        batch(count),
        rows(count, filename=filename),
    )


def test_workbook_has_four_sheets_numeric_scores_and_stable_hundred_item_order() -> None:
    artifact = build_export_workbook(
        snapshot(100),
        now=datetime(2026, 7, 22, 12, 0, tzinfo=UTC),
    )
    workbook = load_workbook(io.BytesIO(artifact.content), data_only=False, keep_links=False)
    assert workbook.sheetnames == ["Summary", "Criteria", "Feedback", "Metadata"]
    assert [cell.value for cell in workbook["Summary"][2]] == [
        "No.",
        "Original filename",
        "Paper status",
        "Result source",
        "Confirmation",
        "Subtotal",
        "Deductions",
        "Final score",
        "Max score",
        "Grade status",
    ]
    assert [cell.value for cell in workbook["Criteria"][1]] == [
        "No.",
        "Original filename",
        "row_type",
        "criterion_id",
        "criterion_name",
        "maximum_or_deduction",
        "score",
        "applied",
        "reason",
        "evidence_block_ids",
    ]
    assert [cell.value for cell in workbook["Feedback"][1]] == [
        "No.",
        "Original filename",
        "Overall feedback",
        "Revision suggestions",
        "Teacher change reason",
        "Result source",
    ]
    assert [cell.value for cell in workbook["Metadata"][1]] == ["Key", "Value"]
    assert workbook["Summary"].max_row == 102
    assert workbook["Summary"].auto_filter.ref == "A2:J102"
    assert workbook["Criteria"].max_row == 201
    assert workbook["Feedback"].max_row == 101
    assert workbook["Summary"]["B1"].value == "非最终成绩"
    assert [workbook["Summary"].cell(row, 1).value for row in range(3, 103)] == list(range(1, 101))
    assert [workbook["Criteria"].cell(row, 1).value for row in range(2, 202)] == [
        paper for paper in range(1, 101) for _ in range(2)
    ]
    assert [workbook["Feedback"].cell(row, 1).value for row in range(2, 102)] == list(range(1, 101))
    assert Decimal(str(workbook["Summary"]["H3"].value)) == Decimal("8.25")
    assert workbook["Criteria"]["H3"].value is True
    assert workbook["Criteria"]["H3"].data_type == "b"
    assert workbook["Criteria"]["H3"].number_format == "General"
    assert artifact.safe_filename.endswith("-33333333-draft-20260722T120000Z.xlsx")
    assert len(artifact.file_sha256) == 32


def test_hundred_unique_papers_keep_criteria_and_feedback_on_the_matching_row() -> None:
    item_rows = list(rows(100))
    for position, item in enumerate(item_rows):
        result = item["result_snapshot"]
        assert isinstance(result, dict)
        criteria = result["criteria_results"]
        assert isinstance(criteria, list)
        criteria[0]["reason"] = f"Reason for paper {position + 1}."
        result["overall_feedback"] = f"Feedback for paper {position + 1}."
    frozen = workbook_snapshot_from_frozen(
        EXPORT_ID,
        "draft",
        "paper-grading-workbook.v1",
        datetime(2026, 7, 22, tzinfo=UTC),
        batch(100),
        tuple(item_rows),
    )
    artifact = build_export_workbook(frozen)
    workbook = load_workbook(io.BytesIO(artifact.content), data_only=False, keep_links=True)

    for position, item in enumerate(item_rows):
        criteria_row = 2 + position * 2
        feedback_row = 2 + position
        assert workbook["Criteria"].cell(criteria_row, 2).value == item["original_filename"]
        assert (
            workbook["Criteria"].cell(criteria_row, 9).value == f"Reason for paper {position + 1}."
        )
        assert workbook["Feedback"].cell(feedback_row, 2).value == item["original_filename"]
        assert (
            workbook["Feedback"].cell(feedback_row, 3).value
            == f"Feedback for paper {position + 1}."
        )


def test_final_workbook_accepts_only_confirmed_review_and_marks_it_final() -> None:
    item_rows = list(rows(1))
    item_rows[0]["source_type"] = "teacher_confirmed"
    item_rows[0]["teacher_review_id"] = UUID("66666666-6666-4666-8666-666666666666")
    item_rows[0]["review_revision"] = 3
    result = item_rows[0]["result_snapshot"]
    assert isinstance(result, dict)
    result["item_status"] = "completed"
    result["confirmed_at"] = "2026-07-22T09:00:00+00:00"
    result["change_reason"] = "Teacher confirmed the evidence."
    frozen = workbook_snapshot_from_frozen(
        EXPORT_ID,
        "final",
        "paper-grading-workbook.v1",
        datetime(2026, 7, 22, tzinfo=UTC),
        batch(1),
        tuple(item_rows),
    )
    artifact = build_export_workbook(frozen)
    workbook = load_workbook(io.BytesIO(artifact.content), data_only=False, keep_links=True)
    metadata = {
        workbook["Metadata"].cell(row, 1).value: workbook["Metadata"].cell(row, 2).value
        for row in range(2, workbook["Metadata"].max_row + 1)
    }

    assert workbook["Summary"]["B1"].value == "教师已确认"
    assert workbook["Summary"]["E3"].value == "confirmed"
    assert workbook["Summary"]["J3"].value == "教师已确认"
    assert metadata["item.1.source_type"] == "teacher_confirmed"
    assert metadata["item.1.review_revision"] == "3"
    assert metadata["item.1.confirmed_at"] == "2026-07-22T09:00:00+00:00"


@pytest.mark.parametrize("prefix", ["=", "+", "-", "@", "\t", "\r", "\n"])
def test_formula_injection_is_literal_text_after_reopen(prefix: str) -> None:
    artifact = build_export_workbook(snapshot(filename=f'{prefix}HYPERLINK("bad")'))
    workbook = load_workbook(io.BytesIO(artifact.content), data_only=False, keep_links=True)
    cell = workbook["Summary"]["B3"]
    assert cell.data_type != "f"
    assert cell.value == f'\'{prefix}HYPERLINK("bad")'
    assert cell.number_format == "@"
    assert not getattr(workbook, "_external_links", [])


def test_all_external_text_categories_share_the_formula_injection_boundary() -> None:
    batch_snapshot = batch(1)
    batch_snapshot["assignment_title"] = "=assignment"
    batch_snapshot["model"] = "\nmodel"
    batch_snapshot["prompt_version"] = "-prompt"
    rubric = batch_snapshot["rubric"]
    assert isinstance(rubric, dict)
    dimensions = rubric["dimensions"]
    assert isinstance(dimensions, list)
    dimensions[0]["name"] = "+rubric"

    item_rows = list(rows(1))
    result = item_rows[0]["result_snapshot"]
    assert isinstance(result, dict)
    criteria = result["criteria_results"]
    assert isinstance(criteria, list)
    criteria[0]["reason"] = "@reason"
    criteria[0]["revision_suggestions"] = ["\trevision"]
    result["overall_feedback"] = "\rfeedback"
    result["change_reason"] = "=teacher change"

    frozen = workbook_snapshot_from_frozen(
        EXPORT_ID,
        "draft",
        "paper-grading-workbook.v1",
        datetime(2026, 7, 22, tzinfo=UTC),
        batch_snapshot,
        tuple(item_rows),
    )
    artifact = build_export_workbook(frozen)
    workbook = load_workbook(io.BytesIO(artifact.content), data_only=False, keep_links=True)
    metadata = {
        workbook["Metadata"].cell(row, 1).value: workbook["Metadata"].cell(row, 2)
        for row in range(2, workbook["Metadata"].max_row + 1)
    }
    checked_cells = (
        workbook["Criteria"]["E2"],
        workbook["Criteria"]["I2"],
        workbook["Feedback"]["C2"],
        workbook["Feedback"]["D2"],
        workbook["Feedback"]["E2"],
        metadata["assignment_title"],
        metadata["model"],
        metadata["prompt_version"],
    )

    assert [cell.value for cell in checked_cells] == [
        "'+rubric",
        "'@reason",
        "'\rfeedback",
        "argument: '\trevision",
        "'=teacher change",
        "'=assignment",
        "'\nmodel",
        "'-prompt",
    ]
    assert all(cell.data_type != "f" and cell.number_format == "@" for cell in checked_cells)
    assert not getattr(workbook, "_external_links", [])


def test_same_frozen_snapshot_produces_identical_bytes_and_hash() -> None:
    first = build_export_workbook(snapshot())
    second = build_export_workbook(snapshot())

    assert second.content == first.content
    assert second.file_sha256 == first.file_sha256
    assert second.safe_filename == first.safe_filename


def test_metadata_values_are_text_even_when_hashes_contain_only_digits() -> None:
    artifact = build_export_workbook(snapshot())
    workbook = load_workbook(io.BytesIO(artifact.content), data_only=False, keep_links=True)
    metadata = workbook["Metadata"]
    values = {
        metadata.cell(row, 1).value: metadata.cell(row, 2) for row in range(2, metadata.max_row + 1)
    }

    assert values["prompt_hash"].value == "22" * 32
    assert values["prompt_hash"].data_type == "s"
    assert values["prompt_hash"].number_format == "@"


def test_metadata_contains_the_required_audit_fields_and_omits_secrets() -> None:
    frozen = snapshot()
    frozen.batch.update(
        {
            "api_key": "never-export",  # pragma: allowlist secret
            "token": "never-export",
            "object_key": "never-export",
            "file_sha256": "never-export",
            "raw_model_response": "never-export",
        }
    )
    artifact = build_export_workbook(frozen)
    workbook = load_workbook(io.BytesIO(artifact.content), data_only=False, keep_links=True)
    metadata = workbook["Metadata"]
    keys = {metadata.cell(row, 1).value for row in range(2, metadata.max_row + 1)}

    assert {
        "export_id",
        "export_type",
        "snapshot_at",
        "generated_at_utc",
        "assignment_id",
        "grading_job_id",
        "assignment_title",
        "rubric_version_id",
        "rubric_version",
        "provider_config_id",
        "provider_config_version",
        "model",
        "prompt_version",
        "prompt_hash",
        "model_parameters_hash",
        "workbook_schema_version",
        "item.1.attempt_id",
        "item.1.review_id",
        "item.1.review_revision",
        "item.1.source_type",
        "item.1.confirmed_at",
    } <= keys
    assert {
        "api_key",
        "token",
        "object_key",
        "file_sha256",
        "raw_model_response",
    }.isdisjoint(keys)


def test_long_assignment_title_uses_a_collision_safe_filename_abbreviation() -> None:
    frozen = snapshot()
    title = "Very long assignment " * 20
    frozen.batch["assignment_title"] = title
    artifact = build_export_workbook(frozen)

    digest = hashlib.sha256("_".join(title.split()).encode()).hexdigest()[:8]
    assert f"-{digest}-33333333-draft-" in artifact.safe_filename
    workbook = load_workbook(io.BytesIO(artifact.content), data_only=False, keep_links=True)
    metadata = workbook["Metadata"]
    values = {
        metadata.cell(row, 1).value: metadata.cell(row, 2).value
        for row in range(2, metadata.max_row + 1)
    }
    assert values["assignment_title"] == title


@pytest.mark.parametrize(
    ("filename", "code"),
    [
        ("bad\x00name.pdf", "export_cell_text_invalid"),
        ("bad\ud800name.pdf", "export_cell_text_invalid"),
        ("bad\ufffename.pdf", "export_cell_text_invalid"),
        ("bad\uffffname.pdf", "export_cell_text_invalid"),
        ("x" * 32768, "export_cell_text_too_long"),
    ],
)
def test_invalid_excel_text_fails_the_whole_export(filename: str, code: str) -> None:
    with pytest.raises(WorkbookValidationError, match=code):
        build_export_workbook(snapshot(filename=filename))


def test_totals_mismatch_fails_instead_of_recalculating_or_truncating() -> None:
    item_rows = list(rows(1))
    result = item_rows[0]["result_snapshot"]
    assert isinstance(result, dict)
    result["final_score"] = "9"
    frozen = workbook_snapshot_from_frozen(
        EXPORT_ID,
        "draft",
        "paper-grading-workbook.v1",
        datetime(2026, 7, 22, tzinfo=UTC),
        batch(1),
        tuple(item_rows),
    )
    with pytest.raises(WorkbookValidationError, match="export_totals_mismatch"):
        build_export_workbook(frozen)


def test_score_outside_rubric_step_fails_the_whole_export() -> None:
    item_rows = list(rows(1))
    result = item_rows[0]["result_snapshot"]
    assert isinstance(result, dict)
    criteria = result["criteria_results"]
    assert isinstance(criteria, list)
    criteria[0]["score"] = "9.10"
    result["subtotal"] = "9.10"
    result["final_score"] = "8.10"
    frozen = workbook_snapshot_from_frozen(
        EXPORT_ID,
        "draft",
        "paper-grading-workbook.v1",
        datetime(2026, 7, 22, tzinfo=UTC),
        batch(1),
        tuple(item_rows),
    )
    with pytest.raises(WorkbookValidationError, match="export_totals_mismatch"):
        build_export_workbook(frozen)


def test_deduction_applied_must_be_a_real_boolean() -> None:
    item_rows = list(rows(1))
    result = item_rows[0]["result_snapshot"]
    assert isinstance(result, dict)
    deductions = result["deduction_results"]
    assert isinstance(deductions, list)
    deductions[0]["applied"] = "false"
    frozen = workbook_snapshot_from_frozen(
        EXPORT_ID,
        "draft",
        "paper-grading-workbook.v1",
        datetime(2026, 7, 22, tzinfo=UTC),
        batch(1),
        tuple(item_rows),
    )

    with pytest.raises(WorkbookValidationError, match="export_snapshot_invalid"):
        build_export_workbook(frozen)


def test_composed_revision_suggestions_cannot_be_silently_truncated() -> None:
    item_rows = list(rows(1))
    result = item_rows[0]["result_snapshot"]
    assert isinstance(result, dict)
    criteria = result["criteria_results"]
    assert isinstance(criteria, list)
    criteria[0]["revision_suggestions"] = ["x" * 18_000, "y" * 18_000]
    frozen = workbook_snapshot_from_frozen(
        EXPORT_ID,
        "draft",
        "paper-grading-workbook.v1",
        datetime(2026, 7, 22, tzinfo=UTC),
        batch(1),
        tuple(item_rows),
    )

    with pytest.raises(WorkbookValidationError, match="export_cell_text_too_long"):
        build_export_workbook(frozen)
