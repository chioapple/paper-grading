"""从不可变导出快照生成无公式、可审计的四工作表 Excel。"""

import hashlib
import io
import re
import unicodedata
import zipfile
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal
from typing import Any, Literal, cast
from uuid import UUID

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.domain.rubric import StructuredRubric

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
SHEET_NAMES = ("Summary", "Criteria", "Feedback", "Metadata")
ILLEGAL_XML = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\ud800-\udfff\ufffe\uffff]")
FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r", "\n")
CORE_MODIFIED = re.compile(rb"(<dcterms:modified\b[^>]*>)[^<]*(</dcterms:modified>)")


class WorkbookValidationError(RuntimeError):
    def __init__(self, code: str) -> None:
        super().__init__(code)
        self.code = code


@dataclass(frozen=True, slots=True)
class WorkbookItem:
    position: int
    submission_id: UUID
    attempt_id: UUID
    review_id: UUID | None
    review_revision: int | None
    source_type: Literal["ai_suggestion", "teacher_draft", "teacher_confirmed"]
    original_filename: str
    result: dict[str, object]


@dataclass(frozen=True, slots=True)
class WorkbookSnapshot:
    export_id: UUID
    export_type: Literal["draft", "final"]
    workbook_schema_version: str
    snapshot_at: datetime
    batch: dict[str, object]
    rubric: StructuredRubric
    items: tuple[WorkbookItem, ...]


@dataclass(frozen=True, slots=True)
class WorkbookArtifact:
    content: bytes
    safe_filename: str
    file_sha256: bytes
    file_size_bytes: int


def _text(value: object) -> str:
    if not isinstance(value, str):
        raise WorkbookValidationError("export_snapshot_invalid")
    if len(value) > 32767:
        raise WorkbookValidationError("export_cell_text_too_long")
    if ILLEGAL_XML.search(value):
        raise WorkbookValidationError("export_cell_text_invalid")
    try:
        value.encode("utf-8")
    except UnicodeEncodeError as error:
        raise WorkbookValidationError("export_cell_text_invalid") from error
    if value.startswith(FORMULA_PREFIXES):
        return "'" + value
    return value


def _decimal(value: object) -> Decimal:
    if isinstance(value, bool) or not isinstance(value, (str, int, float, Decimal)):
        raise WorkbookValidationError("export_score_invalid")
    try:
        result = Decimal(str(value))
    except Exception as error:
        raise WorkbookValidationError("export_score_invalid") from error
    exponent = result.as_tuple().exponent
    if not result.is_finite() or not isinstance(exponent, int) or exponent < -4:
        raise WorkbookValidationError("export_score_invalid")
    return result


def _uuid(value: object) -> UUID:
    try:
        return value if isinstance(value, UUID) else UUID(str(value))
    except (TypeError, ValueError, AttributeError) as error:
        raise WorkbookValidationError("export_snapshot_invalid") from error


def workbook_snapshot_from_frozen(
    export_id: UUID,
    export_type: str,
    workbook_schema_version: str,
    snapshot_at: datetime,
    batch_snapshot: dict[str, object],
    item_rows: tuple[dict[str, object], ...],
) -> WorkbookSnapshot:
    try:
        if export_type not in {"draft", "final"}:
            raise ValueError
        rubric = StructuredRubric.model_validate(batch_snapshot["rubric"])
        items = tuple(
            WorkbookItem(
                position=cast(int, row["position"]),
                submission_id=_uuid(row["submission_id"]),
                attempt_id=_uuid(row["grading_attempt_id"]),
                review_id=(
                    _uuid(row["teacher_review_id"])
                    if row.get("teacher_review_id") is not None
                    else None
                ),
                review_revision=cast(int | None, row.get("review_revision")),
                source_type=cast(
                    Literal["ai_suggestion", "teacher_draft", "teacher_confirmed"],
                    row["source_type"],
                ),
                original_filename=cast(str, row["original_filename"]),
                result=cast(dict[str, object], row["result_snapshot"]),
            )
            for row in item_rows
        )
    except (KeyError, TypeError, ValueError) as error:
        raise WorkbookValidationError("export_snapshot_invalid") from error
    if (
        workbook_schema_version != "paper-grading-workbook.v1"
        or not items
        or len(items) > 100
        or [item.position for item in items] != list(range(len(items)))
        or len({item.submission_id for item in items}) != len(items)
        or batch_snapshot.get("paper_count") != len(items)
    ):
        raise WorkbookValidationError("export_snapshot_invalid")
    return WorkbookSnapshot(
        export_id=export_id,
        export_type=cast(Literal["draft", "final"], export_type),
        workbook_schema_version=workbook_schema_version,
        snapshot_at=snapshot_at,
        batch=batch_snapshot,
        rubric=rubric,
        items=items,
    )


def _validate_result(snapshot: WorkbookSnapshot, item: WorkbookItem) -> dict[str, object]:
    result = item.result
    try:
        criteria_value = result["criteria_results"]
        deductions_value = result["deduction_results"]
        if not isinstance(criteria_value, list) or not all(
            isinstance(row, dict) for row in criteria_value
        ):
            raise TypeError
        if not isinstance(deductions_value, list) or not all(
            isinstance(row, dict) for row in deductions_value
        ):
            raise TypeError
        criteria = cast(list[dict[str, object]], criteria_value)
        deductions = cast(list[dict[str, object]], deductions_value)
        subtotal = _decimal(result["subtotal"])
        deduction_total = _decimal(result["deduction_total"])
        final_score = _decimal(result["final_score"])
        max_score = _decimal(result["max_score"])
    except (KeyError, TypeError) as error:
        raise WorkbookValidationError("export_snapshot_invalid") from error
    dimension_map = {dimension.id: dimension for dimension in snapshot.rubric.dimensions}
    deduction_map = {deduction.id: deduction for deduction in snapshot.rubric.deductions}
    try:
        criterion_ids = [_text(row["dimension_id"]) for row in criteria]
        deduction_ids = [_text(row["deduction_id"]) for row in deductions]
        for row in criteria:
            _text(row["reason"])
            suggestions = row.get("revision_suggestions", [])
            if not isinstance(suggestions, list):
                raise TypeError
            for suggestion in suggestions:
                _text(suggestion)
            _validate_evidence(row.get("evidence", []))
        for row in deductions:
            if not isinstance(row.get("applied"), bool):
                raise TypeError
            _text(row["reason"])
            _validate_evidence(row.get("evidence", []))
        _validate_evidence(result.get("evidence", []), require_target=True)
        _text(result["item_status"])
        _text(result["overall_feedback"])
        if result.get("change_reason") is not None:
            _text(result["change_reason"])
        criterion_scores = {_text(row["dimension_id"]): _decimal(row["score"]) for row in criteria}
        computed_subtotal = sum(criterion_scores.values(), Decimal(0))
        computed_deductions = sum(
            (
                deduction_map[_text(row["deduction_id"])].points
                for row in deductions
                if row.get("applied") is True
            ),
            Decimal(0),
        )
    except (KeyError, TypeError) as error:
        raise WorkbookValidationError("export_snapshot_invalid") from error
    if (
        criterion_ids != list(dimension_map)
        or deduction_ids != list(deduction_map)
        or any(
            score < 0
            or score > dimension_map[dimension_id].max_score
            or score % snapshot.rubric.score_step != 0
            for dimension_id, score in criterion_scores.items()
        )
        or computed_subtotal != subtotal
        or computed_deductions != deduction_total
        or max(Decimal(0), subtotal - deduction_total) != final_score
        or max_score != snapshot.rubric.total_score
    ):
        raise WorkbookValidationError("export_totals_mismatch")
    if snapshot.export_type == "final" and item.source_type != "teacher_confirmed":
        raise WorkbookValidationError("export_final_unconfirmed")
    return result


def _validate_evidence(value: object, *, require_target: bool = False) -> None:
    if not isinstance(value, list) or not all(isinstance(entry, dict) for entry in value):
        raise TypeError
    for raw_entry in value:
        entry = cast(dict[str, object], raw_entry)
        _text(entry["block_id"])
        _text(entry["quote"])
        if require_target:
            _text(entry["target_type"])
            _text(entry["target_id"])


def _numeric(value: Decimal) -> float:
    return float(value)


def _style_sheet(ws: Any, header_row: int = 1) -> None:
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.showGridLines = False
    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, str):
                cell.number_format = "@"
    for column in range(1, ws.max_column + 1):
        values = [
            len(str(ws.cell(row=row, column=column).value or ""))
            for row in range(1, ws.max_row + 1)
        ]
        ws.column_dimensions[get_column_letter(column)].width = min(
            45, max(12, min(max(values, default=12) + 2, 45))
        )
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def _evidence_ids(result: dict[str, object], target_type: str, target_id: str) -> str:
    ids: list[str] = []
    for entry in cast(list[dict[str, object]], result.get("evidence", [])):
        if entry.get("target_type") == target_type and entry.get("target_id") == target_id:
            ids.append(_text(entry.get("block_id")))
    return _text("; ".join(ids))


def _safe_filename(title: str, job_id: UUID, export_type: str, generated_at: datetime) -> str:
    cleaned = unicodedata.normalize("NFKC", _text(title))
    cleaned = re.sub(r"[\\/:*?\"<>|]+", "_", cleaned)
    cleaned = re.sub(r"\s+", "_", cleaned).strip("._") or "assignment"
    if len(cleaned) > 80:
        digest = hashlib.sha256(cleaned.encode("utf-8")).hexdigest()[:8]
        cleaned = f"{cleaned[:71]}-{digest}"
    stamp = generated_at.astimezone(UTC).strftime("%Y%m%dT%H%M%SZ")
    return f"{cleaned}-{str(job_id)[:8]}-{export_type}-{stamp}.xlsx"


def _canonicalize_xlsx(content: bytes, *, workbook_time: datetime) -> bytes:
    """固定 ZIP 成员顺序和时间，保证同一冻结快照可安全重领。"""

    output = io.BytesIO()
    modified_at = workbook_time.isoformat(timespec="seconds").encode("ascii") + b"Z"
    with (
        zipfile.ZipFile(io.BytesIO(content), "r") as source,
        zipfile.ZipFile(
            output,
            "w",
            compression=zipfile.ZIP_DEFLATED,
            compresslevel=9,
        ) as destination,
    ):
        for name in sorted(source.namelist()):
            source_info = source.getinfo(name)
            payload = source.read(name)
            if name == "docProps/core.xml":
                payload, replacements = CORE_MODIFIED.subn(
                    lambda match: match.group(1) + modified_at + match.group(2),
                    payload,
                )
                if replacements != 1:
                    raise WorkbookValidationError("export_workbook_invalid")
            target_info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            target_info.compress_type = zipfile.ZIP_DEFLATED
            target_info.create_system = 0
            target_info.external_attr = 0
            target_info.comment = source_info.comment
            destination.writestr(target_info, payload, compresslevel=9)
    return output.getvalue()


def build_export_workbook(
    snapshot: WorkbookSnapshot, *, now: datetime | None = None
) -> WorkbookArtifact:
    generated_source = now or snapshot.snapshot_at
    if generated_source.tzinfo is None:
        raise WorkbookValidationError("export_snapshot_invalid")
    generated_at = generated_source.astimezone(UTC)
    wb = Workbook()
    workbook_time = generated_at.replace(tzinfo=None)
    wb.properties.created = workbook_time
    wb.properties.modified = workbook_time
    wb.remove(wb.active)
    summary = wb.create_sheet("Summary")
    criteria_sheet = wb.create_sheet("Criteria")
    feedback = wb.create_sheet("Feedback")
    metadata = wb.create_sheet("Metadata")
    numeric_expectations: dict[tuple[str, int, int], Decimal] = {}
    status_label = "教师已确认" if snapshot.export_type == "final" else "非最终成绩"
    summary.append(["导出状态", status_label])
    summary.append(
        [
            "No.",
            "Original filename",
            "Paper status",
            "Result source",
            "Confirmation",
            "Subtotal",
            "Deductions",
            "Final score",
            "Max score",
            "Grade status",
        ]
    )
    criteria_sheet.append(
        [
            "No.",
            "Original filename",
            "row_type",
            "criterion_id",
            "criterion_name",
            "maximum_or_deduction",
            "score",
            "applied",
            "reason",
            "evidence_block_ids",
        ]
    )
    feedback.append(
        [
            "No.",
            "Original filename",
            "Overall feedback",
            "Revision suggestions",
            "Teacher change reason",
            "Result source",
        ]
    )
    dimension_map = {row.id: row for row in snapshot.rubric.dimensions}
    deduction_map = {row.id: row for row in snapshot.rubric.deductions}
    for item in snapshot.items:
        result = _validate_result(snapshot, item)
        criteria = cast(list[dict[str, object]], result["criteria_results"])
        deductions = cast(list[dict[str, object]], result["deduction_results"])
        summary_row = summary.max_row + 1
        summary_values = {
            6: _decimal(result["subtotal"]),
            7: _decimal(result["deduction_total"]),
            8: _decimal(result["final_score"]),
            9: _decimal(result["max_score"]),
        }
        summary.append(
            [
                item.position + 1,
                _text(item.original_filename),
                _text(result["item_status"]),
                item.source_type,
                "confirmed" if item.source_type == "teacher_confirmed" else "not_confirmed",
                _numeric(summary_values[6]),
                _numeric(summary_values[7]),
                _numeric(summary_values[8]),
                _numeric(summary_values[9]),
                status_label,
            ]
        )
        for column, expected in summary_values.items():
            numeric_expectations[("Summary", summary_row, column)] = expected
        suggestions: list[str] = []
        for row in criteria:
            dimension_id = cast(str, row["dimension_id"])
            dimension = dimension_map[dimension_id]
            evidence = row.get("evidence", [])
            evidence_ids = _text(
                "; ".join(
                    _text(cast(dict[str, object], e)["block_id"])
                    for e in cast(list[object], evidence)
                )
            ) or _evidence_ids(result, "dimension", dimension_id)
            row_suggestions = cast(list[str], row.get("revision_suggestions", []))
            suggestions.extend(f"{dimension_id}: {_text(value)}" for value in row_suggestions)
            criteria_row = criteria_sheet.max_row + 1
            criterion_score = _decimal(row["score"])
            criteria_sheet.append(
                [
                    item.position + 1,
                    _text(item.original_filename),
                    "dimension",
                    dimension_id,
                    _text(dimension.name),
                    _numeric(dimension.max_score),
                    _numeric(criterion_score),
                    None,
                    _text(row["reason"]),
                    evidence_ids,
                ]
            )
            numeric_expectations[("Criteria", criteria_row, 6)] = dimension.max_score
            numeric_expectations[("Criteria", criteria_row, 7)] = criterion_score
        for row in deductions:
            deduction_id = cast(str, row["deduction_id"])
            deduction = deduction_map[deduction_id]
            evidence = row.get("evidence", [])
            evidence_ids = _text(
                "; ".join(
                    _text(cast(dict[str, object], e)["block_id"])
                    for e in cast(list[object], evidence)
                )
            ) or _evidence_ids(result, "deduction", deduction_id)
            criteria_row = criteria_sheet.max_row + 1
            criteria_sheet.append(
                [
                    item.position + 1,
                    _text(item.original_filename),
                    "deduction",
                    deduction_id,
                    _text(deduction.name),
                    _numeric(deduction.points),
                    None,
                    row["applied"],
                    _text(row["reason"]),
                    evidence_ids,
                ]
            )
            numeric_expectations[("Criteria", criteria_row, 6)] = deduction.points
        feedback.append(
            [
                item.position + 1,
                _text(item.original_filename),
                _text(result["overall_feedback"]),
                _text("\n".join(suggestions)),
                _text(result["change_reason"]) if result.get("change_reason") is not None else "",
                item.source_type,
            ]
        )
    metadata.append(["Key", "Value"])
    batch_rows = [
        ("export_id", str(snapshot.export_id)),
        ("export_type", snapshot.export_type),
        ("snapshot_at", snapshot.snapshot_at.isoformat()),
        ("generated_at_utc", generated_at.isoformat()),
        ("assignment_id", str(snapshot.batch.get("assignment_id", ""))),
        ("grading_job_id", str(snapshot.batch.get("grading_job_id", ""))),
        ("assignment_title", _text(snapshot.batch["assignment_title"])),
        ("rubric_version_id", str(snapshot.batch["rubric_version_id"])),
        ("rubric_version", str(snapshot.batch["rubric_version"])),
        ("provider_config_id", str(snapshot.batch["provider_config_id"])),
        ("provider_config_version", str(snapshot.batch["provider_config_version"])),
        ("model", _text(snapshot.batch["model"])),
        ("prompt_version", _text(snapshot.batch["prompt_version"])),
        ("prompt_hash", _text(snapshot.batch["prompt_hash"])),
        ("model_parameters_hash", _text(snapshot.batch["model_parameters_hash"])),
        ("workbook_schema_version", snapshot.workbook_schema_version),
    ]
    for metadata_row in batch_rows:
        metadata.append(list(metadata_row))
    for item in snapshot.items:
        result = item.result
        prefix = f"item.{item.position + 1}"
        for key, value in (
            ("attempt_id", item.attempt_id),
            ("review_id", item.review_id or ""),
            ("review_revision", item.review_revision or ""),
            ("source_type", item.source_type),
            ("confirmed_at", result.get("confirmed_at") or ""),
        ):
            metadata.append([f"{prefix}.{key}", _text(str(value))])
    _style_sheet(summary, header_row=2)
    summary["A1"].font = Font(bold=True)
    summary["B1"].font = Font(
        bold=True, color="9C0006" if snapshot.export_type == "draft" else "006100"
    )
    summary["B1"].fill = PatternFill(
        "solid", fgColor="FFC7CE" if snapshot.export_type == "draft" else "C6EFCE"
    )
    for ws in (criteria_sheet, feedback, metadata):
        _style_sheet(ws)
    for cell in criteria_sheet["H"][1:]:
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    for ws in (summary, criteria_sheet):
        for row in ws.iter_rows(min_row=3 if ws is summary else 2):
            for cell in row:
                if (
                    isinstance(cell.value, (int, float))
                    and not isinstance(cell.value, bool)
                    and cell.column >= 6
                ):
                    cell.number_format = "0.####"
    output = io.BytesIO()
    wb.save(output)
    content = _canonicalize_xlsx(output.getvalue(), workbook_time=workbook_time)
    reopened = load_workbook(io.BytesIO(content), data_only=False, keep_links=True)
    if tuple(reopened.sheetnames) != SHEET_NAMES or getattr(reopened, "_external_links", []):
        raise WorkbookValidationError("export_workbook_invalid")
    for ws in reopened.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f":
                    raise WorkbookValidationError("export_formula_detected")
    for (sheet_name, row_number, column), expected in numeric_expectations.items():
        actual = reopened[sheet_name].cell(row=row_number, column=column).value
        if _decimal(actual) != expected:
            raise WorkbookValidationError("export_totals_mismatch")
    filename = _safe_filename(
        cast(str, snapshot.batch["assignment_title"]),
        _uuid(snapshot.batch["grading_job_id"]),
        snapshot.export_type,
        generated_at,
    )
    return WorkbookArtifact(
        content=content,
        safe_filename=filename,
        file_sha256=hashlib.sha256(content).digest(),
        file_size_bytes=len(content),
    )
